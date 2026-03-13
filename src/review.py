# review.py — Automated data review for RDLS metadata preparation
# ----------------------------------------------------------------
# Takes an inventory JSON (from inventory.py) and produces:
# 1. File inspection results (geospatial headers, tabular schemas, doc text)
# 2. HEVL classification per file group
# 3. Gap analysis against RDLS required fields
# 4. Structured review JSON for downstream metadata creation
#
# Dependencies: openpyxl, pandas, geopandas, Pillow (PIL), python-docx
# Optional: netCDF4, xarray (for NetCDF files)
#
# Benny Istanto, GOST/DEC Data Group/The World Bank

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import struct
import sys
import time
import warnings
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .inventory import inventory_folder, InventoryConfig, scan_target, iso_time
from .utils import load_yaml
from .zipaccess import resolve_and_open


# ---------------------------------------------------------------------------
# Knowledge base loader  (configs/review_knowledge.yaml)
# ---------------------------------------------------------------------------

_CONFIGS_DIR = Path(__file__).resolve().parent.parent / "configs"
_review_config: Optional[Dict[str, Any]] = None


def load_review_config(
    yaml_path: Optional[Path] = None,
) -> Dict[str, Any]:
    """Load the review knowledge base from YAML.

    Loads once and caches at module level.  All regex pattern strings
    are compiled to ``re.Pattern`` objects for performance.

    Falls back to built-in defaults if the YAML file is missing so that
    the script never breaks on upgrade.
    """
    global _review_config
    if _review_config is not None and yaml_path is None:
        return _review_config

    if yaml_path is None:
        yaml_path = _CONFIGS_DIR / "review_knowledge.yaml"

    try:
        cfg = load_yaml(yaml_path)
    except FileNotFoundError:
        warnings.warn(
            f"Review knowledge base not found at {yaml_path}; "
            "using built-in defaults.  Create configs/review_knowledge.yaml "
            "for customisation.",
            stacklevel=2,
        )
        cfg = _builtin_defaults()

    compiled = _compile_config(cfg)

    if yaml_path == _CONFIGS_DIR / "review_knowledge.yaml":
        _review_config = compiled

    return compiled


def _compile_config(cfg: Dict[str, Any]) -> Dict[str, Any]:
    """Compile regex strings in the config to ``re.Pattern`` where needed.

    Most patterns are used as raw strings by ``_match_signals`` (via
    ``re.search``), so they stay as strings.  Only patterns that were
    previously pre-compiled in Python (FIAT/HECRAS/MODEL path patterns,
    naming patterns, readme patterns, report patterns, HECRAS dir patterns)
    get compiled here.
    """
    compiled = dict(cfg)  # shallow copy of top-level

    # --- model_software: compile path patterns & dir patterns ---
    ms = compiled.get("model_software", {})
    for model_name, model_info in ms.items():
        if not isinstance(model_info, dict):
            continue
        for key in ("intermediate_path_patterns", "directory_patterns"):
            if key in model_info:
                model_info[f"{key}_compiled"] = [
                    re.compile(p, re.I) for p in model_info[key]
                ]
        # Convert model_extensions list -> set for O(1) lookup
        if "model_extensions" in model_info:
            model_info["model_extensions_set"] = set(model_info["model_extensions"])
        # Convert intermediate_filenames list -> set
        if "intermediate_filenames" in model_info:
            model_info["intermediate_filenames_set"] = set(model_info["intermediate_filenames"])

    # report_folder_patterns lives directly under model_software
    if "report_folder_patterns" in ms:
        ms["report_folder_patterns_compiled"] = [
            re.compile(p, re.I) for p in ms["report_folder_patterns"]
        ]

    # --- naming_patterns: compile regex patterns ---
    np_section = compiled.get("naming_patterns", {})
    if "patterns" in np_section:
        np_section["patterns_compiled"] = [
            re.compile(entry["pattern"], re.I)
            for entry in np_section["patterns"]
        ]

    # --- readme_patterns: compile each pattern ---
    rp = compiled.get("readme_patterns", {})
    for key, info in rp.items():
        if isinstance(info, dict) and "pattern" in info:
            info["compiled"] = re.compile(info["pattern"], re.I)

    # --- column_detection: compile patterns ---
    cd = compiled.get("column_detection", {})
    for group_name, group_info in cd.items():
        if isinstance(group_info, dict) and "patterns" in group_info:
            group_info["patterns_compiled"] = [
                re.compile(p, re.I) for p in group_info["patterns"]
            ]

    return compiled


def _builtin_defaults() -> Dict[str, Any]:
    """Minimal built-in defaults (fallback when YAML file is missing)."""
    return {
        "hevl_signals": {
            "hazard": {}, "exposure": {},
            "vulnerability": {"patterns": []},
            "loss": {"patterns": []},
            "dem": {"patterns": []},
        },
        "file_filtering": {"exclusion_patterns": [], "irrelevant_patterns": []},
        "model_software": {},
        "naming_patterns": {
            "patterns": [], "scenario_map": {},
            "jba_hazard_map": {}, "hazard_subtype_map": {},
        },
        "readme_patterns": {},
        "rdls_schema": {"required": [], "recommended": []},
        "column_detection": {},
    }


def _get_config() -> Dict[str, Any]:
    """Get or load the review config (cached at module level)."""
    global _review_config
    if _review_config is None:
        _review_config = load_review_config()
    return _review_config


# ---------------------------------------------------------------------------
# Module-level constant aliases (loaded from YAML, same names as before)
# ---------------------------------------------------------------------------

_cfg_init = _get_config()

# HEVL signal dictionaries
HAZARD_SIGNALS = _cfg_init["hevl_signals"]["hazard"]
EXPOSURE_SIGNALS = _cfg_init["hevl_signals"]["exposure"]
VULNERABILITY_SIGNALS = _cfg_init["hevl_signals"]["vulnerability"]
LOSS_SIGNALS = _cfg_init["hevl_signals"]["loss"]
DEM_SIGNALS = _cfg_init["hevl_signals"]["dem"]["patterns"]

# File filtering
EXCLUSION_PATTERNS = _cfg_init["file_filtering"]["exclusion_patterns"]

# Model software (pre-compiled)
_ms = _cfg_init["model_software"]
FIAT_INTERMEDIATE_PATTERNS = _ms.get("fiat", {}).get("intermediate_path_patterns_compiled", [])
FIAT_INTERMEDIATE_FILENAMES = _ms.get("fiat", {}).get("intermediate_filenames_set", set())
HECRAS_MODEL_EXTENSIONS = _ms.get("hecras", {}).get("model_extensions_set", set())
HECRAS_DIR_PATTERNS = _ms.get("hecras", {}).get("directory_patterns_compiled", [])
MODEL_INTERMEDIATE_PATTERNS = _ms.get("general", {}).get("intermediate_path_patterns_compiled", [])
REPORT_FOLDER_PATTERNS = _ms.get("report_folder_patterns_compiled", [])

del _ms  # cleanup

# Irrelevant file patterns
IRRELEVANT_PATTERNS = _cfg_init["file_filtering"]["irrelevant_patterns"]

# Naming patterns (pre-compiled)
_np = _cfg_init["naming_patterns"]
_NAMING_PATTERNS = _np.get("patterns_compiled", [])
_SCENARIO_MAP = _np.get("scenario_map", {})
_JBA_HAZARD_MAP = _np.get("jba_hazard_map", {})
_HAZARD_SUBTYPE_MAP = _np.get("hazard_subtype_map", {})
del _np

# README patterns (pre-compiled)
_README_PATTERNS = {
    k: v["compiled"]
    for k, v in _cfg_init.get("readme_patterns", {}).items()
    if isinstance(v, dict) and "compiled" in v
}

# RDLS schema fields
RDLS_REQUIRED = _cfg_init["rdls_schema"]["required"]
RDLS_RECOMMENDED = _cfg_init["rdls_schema"]["recommended"]

# Column-level detection (v2, compiled patterns)
COLUMN_DETECTION = _cfg_init.get("column_detection", {})

# Weight mapping for column signals
_WEIGHT_MAP = {"high": 3, "medium": 2, "low": 1}

del _cfg_init  # cleanup init-only reference


def classify_intermediate_files(
    files: List[str],
) -> Tuple[List[str], List[str], List[str]]:
    """Separate relevant files from intermediate/model/working outputs.

    Handles FIAT per-RP outputs, HEC-RAS model files, backups, lock files,
    and report/documentation files.

    Returns (relevant, intermediate, reasons) where reasons describes
    why files were classified as intermediate.
    """
    relevant: List[str] = []
    intermediate: List[str] = []
    reasons: List[str] = []
    reason_counts: Dict[str, int] = {}

    for fpath in files:
        inner = fpath.split("::")[-1] if "::" in fpath else fpath
        fname = Path(inner).name.lower()
        ext = Path(inner).suffix.lower()

        # Check if file is inside a HEC-RAS project directory
        in_hecras_dir = any(p.search(fpath) for p in HECRAS_DIR_PATTERNS)
        if in_hecras_dir:
            intermediate.append(fpath)
            reason_counts.setdefault("HEC-RAS project directory", 0)
            reason_counts["HEC-RAS project directory"] += 1
            continue

        # Check FIAT filename match
        if fname in FIAT_INTERMEDIATE_FILENAMES:
            intermediate.append(fpath)
            reason_counts.setdefault("FIAT config/summary files", 0)
            reason_counts["FIAT config/summary files"] += 1
            continue

        # Check HEC-RAS model file extensions
        # (but NOT .prj inside a shapefile group — check for accompanying .shp)
        if ext in HECRAS_MODEL_EXTENSIONS:
            # .prj could be a shapefile projection; only classify as model if
            # it's in a Model/ folder or there's no .shp sibling
            if ext == ".prj":
                in_model_dir = bool(re.search(r"[/\\]Model[/\\]", fpath, re.I))
                if in_model_dir:
                    intermediate.append(fpath)
                    reason_counts.setdefault("HEC-RAS model files", 0)
                    reason_counts["HEC-RAS model files"] += 1
                    continue
                else:
                    relevant.append(fpath)
                    continue
            intermediate.append(fpath)
            reason_counts.setdefault("HEC-RAS model files", 0)
            reason_counts["HEC-RAS model files"] += 1
            continue

        # Check FIAT path patterns
        hit = False
        for pat in FIAT_INTERMEDIATE_PATTERNS:
            if pat.search(fpath):
                intermediate.append(fpath)
                reason_counts.setdefault(f"Pattern: {pat.pattern}", 0)
                reason_counts[f"Pattern: {pat.pattern}"] += 1
                hit = True
                break

        # Check model intermediate patterns (lock files, backups, etc.)
        if not hit:
            for pat in MODEL_INTERMEDIATE_PATTERNS:
                if pat.search(fpath):
                    intermediate.append(fpath)
                    reason_counts.setdefault("Model working/lock files", 0)
                    reason_counts["Model working/lock files"] += 1
                    hit = True
                    break

        # Check report/documentation folders
        if not hit:
            for pat in REPORT_FOLDER_PATTERNS:
                if pat.search(fpath):
                    intermediate.append(fpath)
                    reason_counts.setdefault("Report/documentation files", 0)
                    reason_counts["Report/documentation files"] += 1
                    hit = True
                    break

        if not hit:
            relevant.append(fpath)

    for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
        reasons.append(f"{reason} ({count} files)")

    return relevant, intermediate, reasons


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class FileInspection:
    """Result of inspecting a single file's content/metadata."""
    path: str
    format: str  # geotiff, geojson, shapefile, xlsx, csv, json, pdf, docx, etc.
    inspection: Dict[str, Any] = field(default_factory=dict)
    # Inspection can contain: crs, bounds, resolution, bands, columns, schema,
    # feature_count, row_count, sheet_names, text_excerpt, etc.


@dataclass
class FileGroup:
    """A logical grouping of files that form one dataset."""
    name: str
    files: List[str]  # relative paths
    formats: List[str]
    total_size_bytes: int
    hevl: List[str]  # ["H"], ["E"], ["V"], ["L"], or combinations
    hazard_types: List[str] = field(default_factory=list)
    exposure_categories: List[str] = field(default_factory=list)
    confidence: str = "low"  # high, medium, low
    evidence: List[str] = field(default_factory=list)
    inspections: List[FileInspection] = field(default_factory=list)
    conflicts: List[str] = field(default_factory=list)
    column_evidence: List[str] = field(default_factory=list)


@dataclass
class GapAnalysis:
    """Missing or incomplete RDLS metadata fields for a file group."""
    group_name: str
    available: Dict[str, str] = field(default_factory=dict)
    missing_required: List[str] = field(default_factory=list)
    missing_recommended: List[str] = field(default_factory=list)
    actions: List[str] = field(default_factory=list)


@dataclass
class ReviewResult:
    """Complete review output for a folder."""
    target: str
    generated_utc: str
    stats: Dict[str, Any]
    file_groups: List[FileGroup]
    inspections: List[FileInspection]
    gap_analyses: List[GapAnalysis]
    suggested_datasets: List[Dict[str, Any]]
    quality_issues: Dict[str, List[str]] = field(default_factory=dict)
    project_metadata: Dict[str, str] = field(default_factory=dict)
    intermediate_files: Dict[str, Any] = field(default_factory=dict)
    inventory_rows: List[Dict] = field(default_factory=list)


# ---------------------------------------------------------------------------
# File inspection functions
# ---------------------------------------------------------------------------

def inspect_geotiff(path: Path) -> FileInspection:
    """Extract metadata from a GeoTIFF. Prefers rasterio; falls back to PIL."""
    result: Dict[str, Any] = {}

    # Try rasterio first (full metadata)
    try:
        import rasterio
        with rasterio.open(path) as ds:
            result["width"] = ds.width
            result["height"] = ds.height
            result["crs"] = str(ds.crs) if ds.crs else None
            result["crs_epsg"] = ds.crs.to_epsg() if ds.crs else None
            result["bounds"] = list(ds.bounds)  # [left, bottom, right, top]
            result["transform"] = list(ds.transform)[:6]
            result["pixel_scale_x"] = abs(ds.transform[0])
            result["pixel_scale_y"] = abs(ds.transform[4])
            result["band_count"] = ds.count
            result["dtype"] = str(ds.dtypes[0])
            result["nodata"] = ds.nodata
            result["driver"] = ds.driver
            # Read band statistics if available in tags
            tags = ds.tags()
            if tags:
                result["tags"] = {k: v for k, v in list(tags.items())[:10]}

            # Band statistics via center-window sampling (512x512 max)
            try:
                import numpy as np
                band_stats = []
                win_w = min(512, ds.width)
                win_h = min(512, ds.height)
                col_off = (ds.width - win_w) // 2
                row_off = (ds.height - win_h) // 2
                window = rasterio.windows.Window(col_off, row_off, win_w, win_h)
                for band_idx in range(1, min(ds.count + 1, 5)):  # cap at 4 bands
                    data = ds.read(band_idx, window=window)
                    nodata_val = ds.nodata
                    if nodata_val is not None:
                        mask = data != nodata_val
                        valid = data[mask]
                    else:
                        valid = data.ravel()
                    if len(valid) > 0:
                        band_stats.append({
                            "band": band_idx,
                            "min": float(np.nanmin(valid)),
                            "max": float(np.nanmax(valid)),
                            "mean": float(np.nanmean(valid)),
                            "nodata_pct": round(
                                100.0 * (1.0 - len(valid) / data.size), 1
                            ) if nodata_val is not None else 0.0,
                        })
                    else:
                        band_stats.append({
                            "band": band_idx,
                            "note": "all nodata in sample window",
                            "nodata_pct": 100.0,
                        })
                if band_stats:
                    result["band_stats"] = band_stats
            except Exception:
                pass  # stats are best-effort

        return FileInspection(path=str(path), format="geotiff", inspection=result)
    except ImportError:
        pass  # Fall through to PIL

    # Fallback: PIL (basic metadata only)
    try:
        from PIL import Image
        img = Image.open(path)
        result["width"] = img.size[0]
        result["height"] = img.size[1]
        result["mode"] = img.mode
        result["dtype"] = "float32" if img.mode == "F" else img.mode
        result["_inspector"] = "PIL (rasterio not available)"

        tags = img.tag_v2 if hasattr(img, "tag_v2") else {}

        if 33550 in tags:
            scale = tags[33550]
            result["pixel_scale_x"] = scale[0]
            result["pixel_scale_y"] = scale[1]
        if 33922 in tags:
            tp = tags[33922]
            if len(tp) >= 6:
                result["origin_x"] = tp[3]
                result["origin_y"] = tp[4]
        if 34737 in tags:
            crs_str = tags[34737]
            result["crs_string"] = crs_str
            if "UTM" in crs_str:
                m = re.search(r"UTM.*?Zone.*?(\d+[NS]?)", crs_str)
                if m:
                    result["crs"] = f"WGS 84 / UTM Zone {m.group(1)}"
            elif "EPSG" in crs_str:
                m = re.search(r"EPSG[:\s]*(\d+)", crs_str)
                if m:
                    result["crs"] = f"EPSG:{m.group(1)}"
        if "origin_x" in result and "pixel_scale_x" in result:
            result["bounds"] = [
                result["origin_x"],
                result["origin_y"] - result["pixel_scale_y"] * result["height"],
                result["origin_x"] + result["pixel_scale_x"] * result["width"],
                result["origin_y"],
            ]
        img.close()
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="geotiff", inspection=result)


def inspect_vector(path: Path) -> FileInspection:
    """Extract schema + attribute analysis from a vector file.

    Uses fiona for fast feature count (O(1) for shapefiles via .shx header),
    then geopandas for a 200-row sample with column value analysis.
    """
    result: Dict[str, Any] = {}

    # Fast feature count via fiona (does not read all features)
    try:
        import fiona
        with fiona.open(str(path)) as src:
            result["feature_count"] = len(src)
            result["crs"] = str(src.crs) if src.crs else None
            result["driver"] = src.driver
            result["bounds"] = [float(v) for v in src.bounds]
    except Exception:
        pass  # fall through to geopandas

    # Read up to 200 rows for attribute analysis
    try:
        import geopandas as gpd
        import numpy as np
        gdf = gpd.read_file(path, rows=200)
        if "crs" not in result:
            result["crs"] = str(gdf.crs) if gdf.crs else None
        if "bounds" not in result:
            result["bounds"] = [float(v) for v in gdf.total_bounds]
        if "feature_count" not in result:
            result["feature_count"] = len(gdf)  # at least the sample size
        result["geometry_type"] = gdf.geometry.geom_type.iloc[0] if len(gdf) > 0 else None
        result["columns"] = [c for c in gdf.columns if c != "geometry"]
        result["dtypes"] = {c: str(gdf[c].dtype) for c in result["columns"]}
        result["sample_size"] = len(gdf)

        # Column value analysis
        col_details: Dict[str, Dict] = {}
        for col_name in result["columns"]:
            col = gdf[col_name]
            info: Dict[str, Any] = {"dtype": str(col.dtype)}
            try:
                n_unique = int(col.nunique())
                info["unique_count"] = n_unique

                if col.dtype in ("float64", "float32", "int64", "int32", "int16"):
                    # Numeric: show range
                    info["min"] = col.min()
                    info["max"] = col.max()
                    # Convert numpy types to Python native for JSON serialization
                    if hasattr(info["min"], "item"):
                        info["min"] = info["min"].item()
                    if hasattr(info["max"], "item"):
                        info["max"] = info["max"].item()
                elif n_unique <= 30:
                    # Categorical with few values: list them
                    vals = col.dropna().unique().tolist()
                    info["unique_values"] = [
                        v.item() if hasattr(v, "item") else v
                        for v in sorted(vals, key=str)[:30]
                    ]
            except Exception:
                pass
            col_details[col_name] = info
        result["col_details"] = col_details

    except Exception as e:
        if "error" not in result:
            result["error"] = str(e)

    ext = path.suffix.lower()
    fmt = {".geojson": "geojson", ".shp": "shapefile", ".gpkg": "geopackage"}.get(ext, "vector")
    return FileInspection(path=str(path), format=fmt, inspection=result)


def inspect_fgdb(path: Path) -> FileInspection:
    """Inspect ESRI File Geodatabase layers using geopandas/fiona."""
    result: Dict[str, Any] = {}
    try:
        import fiona
        layers = fiona.listlayers(str(path))
        result["layers"] = layers
        result["layer_count"] = len(layers)
        # Inspect first layer
        if layers:
            import geopandas as gpd
            gdf = gpd.read_file(path, layer=layers[0], rows=5)
            result["first_layer"] = layers[0]
            result["crs"] = str(gdf.crs) if gdf.crs else None
            result["columns"] = [c for c in gdf.columns if c != "geometry"]
            result["feature_count_sample"] = len(gdf)
    except ImportError:
        result["note"] = "fiona not available for FGDB reading"
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="fgdb", inspection=result)


def inspect_xlsx(path: Path) -> FileInspection:
    """Read Excel file: sheet names, column headers, row counts, damage-state detection."""
    result: Dict[str, Any] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows_data = list(ws.iter_rows(max_row=5, values_only=True))
            info: Dict[str, Any] = {}
            if rows_data:
                info["columns"] = [str(c) if c is not None else "" for c in rows_data[0]]
                info["col_count"] = len(info["columns"])
                if len(rows_data) > 1:
                    info["sample_row"] = [
                        v if isinstance(v, (str, int, float, type(None))) else str(v)
                        for v in rows_data[1]
                    ]
            # Try to get row count
            try:
                info["max_row"] = ws.max_row
            except Exception:
                pass
            sheets[name] = info
        wb.close()

        # Enhanced: accurate row counts + damage state detection via pandas
        try:
            import pandas as pd
            for sname in list(sheets.keys())[:5]:  # limit to 5 sheets
                try:
                    df = pd.read_excel(path, sheet_name=sname, nrows=10000)
                    sheets[sname]["row_count"] = len(df)

                    # Detect damage state columns (DS1-DS4, muds, sigmads, hw0-hw6)
                    col_lower = [str(c).lower() for c in df.columns]
                    ds_cols = [c for c in df.columns
                               if re.match(r"(?:mu)?ds\d|sigma_?ds\d|hw\d", str(c).lower())]
                    if ds_cols:
                        sheets[sname]["damage_states"] = [str(c) for c in ds_cols]

                    # Unique value counts for first few columns (typology detection)
                    uniques = {}
                    for col in list(df.columns)[:6]:
                        try:
                            uniques[str(col)] = int(df[col].nunique())
                        except Exception:
                            pass
                    if uniques:
                        sheets[sname]["unique_counts"] = uniques

                except Exception:
                    pass
        except ImportError:
            pass

        result["sheets"] = sheets
        result["sheet_count"] = len(sheets)
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="xlsx", inspection=result)


def inspect_csv(path: Path, max_rows: int = 5) -> FileInspection:
    """Read CSV file: headers, sample rows, row count estimate."""
    result: Dict[str, Any] = {}
    try:
        import pandas as pd
        df = pd.read_csv(path, nrows=max_rows)
        result["columns"] = list(df.columns)
        result["dtypes"] = {c: str(df[c].dtype) for c in df.columns}
        result["col_count"] = len(df.columns)
        result["sample_rows"] = max_rows
        if len(df) > 0:
            result["sample_row"] = df.iloc[0].to_dict()
        # Count total rows (fast)
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            result["row_count"] = sum(1 for _ in f) - 1
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="csv", inspection=result)


def inspect_json_data(path: Path) -> FileInspection:
    """Inspect a JSON file: detect structure (array, dict), field names, count."""
    result: Dict[str, Any] = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            result["type"] = "array"
            result["count"] = len(data)
            if data and isinstance(data[0], dict):
                result["fields"] = list(data[0].keys())
                result["sample"] = {
                    k: v for k, v in list(data[0].items())[:8]
                    if isinstance(v, (str, int, float, bool, type(None)))
                }
        elif isinstance(data, dict):
            result["type"] = "object"
            result["keys"] = list(data.keys())[:20]
            # Check if it's a GeoJSON FeatureCollection
            if data.get("type") == "FeatureCollection":
                result["is_geojson"] = True
                result["feature_count"] = len(data.get("features", []))
                if data.get("features"):
                    props = data["features"][0].get("properties", {})
                    result["fields"] = list(props.keys())
            else:
                # Detect columnar JSON: either dict of equal-length arrays
                # OR dict of equal-length dicts (pandas to_json default format)
                col_lens = []
                col_type = None
                for v in data.values():
                    if isinstance(v, list):
                        col_lens.append(len(v))
                        col_type = col_type or "list"
                    elif isinstance(v, dict):
                        col_lens.append(len(v))
                        col_type = col_type or "dict"
                if (col_lens
                        and len(col_lens) == len(data)  # all values are list/dict
                        and len(set(col_lens)) == 1
                        and col_lens[0] > 0):
                    result["is_columnar"] = True
                    result["record_count"] = col_lens[0]
                    result["fields"] = list(data.keys())
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="json", inspection=result)


def inspect_text(path: Path, max_chars: int = 2000) -> FileInspection:
    """Read plain text file (README, txt) up to max_chars."""
    result: Dict[str, Any] = {}
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read(max_chars)
        result["text_excerpt"] = text
        result["char_count"] = len(text)
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="text", inspection=result)


def inspect_netcdf(path: Path) -> FileInspection:
    """Inspect NetCDF file: dimensions, variables, global attributes."""
    result: Dict[str, Any] = {}
    try:
        import netCDF4
        ds = netCDF4.Dataset(str(path), "r")
        result["dimensions"] = {name: len(dim) for name, dim in ds.dimensions.items()}
        result["variables"] = list(ds.variables.keys())
        result["global_attrs"] = {k: str(ds.getncattr(k)) for k in ds.ncattrs()}
        ds.close()
    except ImportError:
        result["note"] = "netCDF4 not available"
    except Exception as e:
        result["error"] = str(e)

    return FileInspection(path=str(path), format="netcdf", inspection=result)


def inspect_file(path: Path) -> Optional[FileInspection]:
    """Dispatch to the appropriate inspector based on file extension."""
    ext = path.suffix.lower()

    if ext in (".tif", ".tiff"):
        return inspect_geotiff(path)
    elif ext == ".geojson":
        # Could be vector or just JSON; try vector first
        return inspect_vector(path)
    elif ext == ".shp":
        return inspect_vector(path)
    elif ext == ".gpkg":
        return inspect_vector(path)
    elif ext == ".xlsx":
        return inspect_xlsx(path)
    elif ext == ".xls":
        # openpyxl doesn't handle .xls; skip or note
        return FileInspection(path=str(path), format="xls",
                              inspection={"note": "Legacy .xls format; use xlrd for reading"})
    elif ext == ".csv":
        return inspect_csv(path)
    elif ext == ".json":
        return inspect_json_data(path)
    elif ext in (".txt", ".md", ".readme"):
        return inspect_text(path)
    elif ext in (".nc", ".nc4", ".hdf5"):
        return inspect_netcdf(path)
    elif ext in (".pdf",):
        return _inspect_pdf(path)
    elif ext in (".docx",):
        return _inspect_docx(path)
    else:
        return None  # Skip unrecognized formats


def _inspect_pdf(path: Path) -> FileInspection:
    """Extract text from PDF using PyMuPDF (fitz)."""
    result: Dict[str, Any] = {}
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(str(path))
        result["page_count"] = len(doc)
        # Extract text from first 5 pages
        pages_text = []
        for i, page in enumerate(doc):
            if i >= 5:
                break
            pages_text.append(page.get_text())
        result["text_excerpt"] = "\n".join(pages_text)[:3000]
        doc.close()
    except ImportError:
        result["note"] = "PyMuPDF (fitz) not available for PDF reading"
    except Exception as e:
        result["error"] = str(e)
    return FileInspection(path=str(path), format="pdf", inspection=result)


def _inspect_docx(path: Path) -> FileInspection:
    """Extract text from DOCX."""
    result: Dict[str, Any] = {}
    try:
        from docx import Document
        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        result["paragraph_count"] = len(paragraphs)
        result["text_excerpt"] = "\n".join(paragraphs[:20])
    except Exception as e:
        result["error"] = str(e)
    return FileInspection(path=str(path), format="docx", inspection=result)


# ---------------------------------------------------------------------------
# ZIP folder tree view (like 7-Zip)
# ---------------------------------------------------------------------------

def build_zip_tree(rows: List[Dict]) -> Dict[str, Any]:
    """Build a hierarchical folder tree from inventory rows for ZIP archives.

    Returns a dict of {zip_name: tree_node} where tree_node has:
      - "files": int (file count at this level)
      - "size": int (total bytes at this level)
      - "children": {name: tree_node}
    """
    zips: Dict[str, Dict] = {}

    for row in rows:
        if not row.get("is_in_zip"):
            continue
        container = row["container"]
        member_path = row.get("path", "")
        # member_path is like "01 Flood hazard layers.zip::subdir/file.tif"
        if "::" in member_path:
            inner = member_path.split("::", 1)[1]
        else:
            continue

        if container not in zips:
            zips[container] = {"files": 0, "size": 0, "children": {}}

        parts = inner.replace("\\", "/").split("/")
        node = zips[container]
        # Navigate/create the tree path
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                # leaf file — count it at current node
                node["files"] += 1
                node["size"] += row.get("size_bytes", 0)
            else:
                if part not in node["children"]:
                    node["children"][part] = {"files": 0, "size": 0, "children": {}}
                node = node["children"][part]

    return zips


def _render_tree(node: Dict, prefix: str = "", is_last: bool = True) -> List[str]:
    """Render a tree node as lines of text (7-Zip style)."""
    lines = []
    children = sorted(node["children"].items())

    for i, (name, child) in enumerate(children):
        last = (i == len(children) - 1)
        connector = "└── " if last else "├── "
        # Count total files in subtree
        total = _count_tree_files(child)
        from .inventory import human_size
        size_str = human_size(child["size"] + sum(
            _sum_tree_size(c) for c in child["children"].values()
        ))
        lines.append(f"{prefix}{connector}{name}/ ({total} files, {size_str})")
        ext_prefix = prefix + ("    " if last else "│   ")
        lines.extend(_render_tree(child, ext_prefix, last))

    return lines


def _count_tree_files(node: Dict) -> int:
    """Count total files in a tree node recursively."""
    total = node["files"]
    for child in node["children"].values():
        total += _count_tree_files(child)
    return total


def _sum_tree_size(node: Dict) -> int:
    """Sum total bytes in a tree node recursively."""
    total = node["size"]
    for child in node["children"].values():
        total += _sum_tree_size(child)
    return total


# ---------------------------------------------------------------------------
# File grouping logic
# ---------------------------------------------------------------------------

def _is_sidecar(name: str) -> bool:
    """Check if a file is a sidecar/companion file (not a primary data file)."""
    return any(re.search(p, name.lower()) for p in EXCLUSION_PATTERNS)


def check_data_quality(group: "FileGroup") -> List[str]:
    """Detect common data quality issues in a file group.

    Returns a list of human-readable issue strings.
    """
    issues: List[str] = []

    # macOS artifact detection
    macos_files = [f for f in group.files if "__MACOSX" in f or ".DS_Store" in f]
    if macos_files:
        issues.append(f"macOS artifacts: {len(macos_files)} files (__MACOSX, .DS_Store)")

    # CRS consistency check
    crs_values = set()
    for insp in group.inspections:
        crs = insp.inspection.get("crs") or insp.inspection.get("crs_short")
        if crs:
            crs_values.add(str(crs))
    if len(crs_values) > 1:
        issues.append(f"CRS inconsistency: {', '.join(sorted(crs_values))}")

    # Sidecar files count
    sidecar_count = sum(1 for f in group.files
                        if any(re.search(p, f, re.I) for p in IRRELEVANT_PATTERNS))
    if sidecar_count > 0:
        issues.append(
            f"Sidecar/metadata files: {sidecar_count}"
            f" (.aux.xml, .ovr, .tfw, etc.)"
        )

    # Empty files
    empty = [f for f in group.files
             if any(i.path == f and i.inspection.get("size_bytes", 1) == 0
                    for i in group.inspections)]
    if empty:
        issues.append(f"Empty files: {len(empty)}")

    return issues


# ---------------------------------------------------------------------------
# README / text metadata extraction
# ---------------------------------------------------------------------------

def extract_readme_metadata(inspections: List["FileInspection"]) -> Dict[str, str]:
    """Extract project metadata from README/text file inspections.

    Looks for project title, provider, contact, date range, etc.
    in text_excerpt fields of text-format inspections.
    """
    metadata: Dict[str, str] = {}
    for insp in inspections:
        if insp.format != "text":
            continue
        text = insp.inspection.get("text_excerpt", "")
        if not text:
            continue
        for key, pattern in _README_PATTERNS.items():
            if key in metadata:
                continue  # keep first match
            m = pattern.search(text)
            if m:
                # For date_range, combine both groups
                if key == "date_range":
                    metadata[key] = f"{m.group(1)} – {m.group(2)}"
                else:
                    metadata[key] = m.group(1).strip()
    return metadata


# ---------------------------------------------------------------------------
# Naming pattern analysis — extract scenarios, return periods, hazard subtypes
# ---------------------------------------------------------------------------

def analyze_naming_patterns(filenames: List[str]) -> Dict[str, Any]:
    """Extract scenarios, return periods, hazard subtypes, GMPEs from file naming.

    Returns a dict with keys: scenarios, return_periods, hazard_subtypes,
    gmpes, intensity_measures, asset_types, matched_count, total_count.
    """
    scenarios: set = set()
    return_periods: set = set()
    hazard_subtypes: set = set()
    gmpes: set = set()
    intensity_measures: set = set()
    asset_types: set = set()
    matched = 0

    for fname in filenames:
        hit = False
        for pat in _NAMING_PATTERNS:
            m = pat.search(fname)
            if m:
                d = m.groupdict()
                if "scenario" in d and d["scenario"]:
                    raw = d["scenario"].upper()
                    scenarios.add(_SCENARIO_MAP.get(raw, raw))
                if "return_period" in d and d["return_period"]:
                    return_periods.add(int(d["return_period"]))
                if "hazard" in d and d["hazard"]:
                    code = d["hazard"].upper()
                    mapped = _JBA_HAZARD_MAP.get(code)
                    if mapped:
                        hazard_subtypes.add(mapped)
                    else:
                        hazard_subtypes.add(
                            _HAZARD_SUBTYPE_MAP.get(code.lower(), code.lower())
                        )
                if "gmpe" in d and d["gmpe"]:
                    gmpes.add(d["gmpe"])
                if "im" in d and d["im"]:
                    intensity_measures.add(d["im"])
                if "asset_type" in d and d["asset_type"]:
                    asset_types.add(d["asset_type"].lower())
                hit = True
        if hit:
            matched += 1

    return {
        "scenarios": sorted(scenarios),
        "return_periods": sorted(return_periods),
        "hazard_subtypes": sorted(hazard_subtypes),
        "gmpes": sorted(gmpes),
        "intensity_measures": sorted(intensity_measures),
        "asset_types": sorted(asset_types),
        "matched_count": matched,
        "total_count": len(filenames),
    }


def group_files(rows: List[Dict], base_path: Path) -> List[FileGroup]:
    """
    Group inventory rows into logical datasets based on folder structure and naming.

    Strategy:
    - Top-level folders become groups
    - For ZIP-based inventories, each ZIP becomes a group
    - Sub-folders within a group can indicate sub-datasets
    """
    groups: Dict[str, Dict] = {}
    sep = os.sep

    for row in rows:
        # Skip inventory output and sidecar files
        if _is_sidecar(row["name"]) or "_inventory" in row["path"] or "_rdls_review" in row["path"]:
            continue

        # Determine group key
        if row.get("is_in_zip"):
            # Group by container (ZIP name)
            group_key = row["container"]
        else:
            path = row["path"]
            parts = path.split(sep)
            if len(parts) > 1:
                group_key = parts[0]
            else:
                group_key = "_root"

        if group_key not in groups:
            groups[group_key] = {
                "files": [],
                "formats": set(),
                "total_bytes": 0,
            }

        g = groups[group_key]
        g["files"].append(row["path"])
        ext = row["ext"].lower().lstrip(".")
        if ext and not _is_sidecar(row["name"]):
            g["formats"].add(ext)
        g["total_bytes"] += row.get("size_bytes", 0)

    result = []
    for key, data in sorted(groups.items()):
        result.append(FileGroup(
            name=key,
            files=data["files"],
            formats=sorted(data["formats"]),
            total_size_bytes=data["total_bytes"],
            hevl=[],
        ))

    return result


# ---------------------------------------------------------------------------
# HEVL classification
# ---------------------------------------------------------------------------

def _match_signals(text: str, patterns: List[str]) -> List[str]:
    """Return list of matched pattern strings."""
    text_lower = text.lower()
    matches = []
    for p in patterns:
        if re.search(p, text_lower):
            matches.append(p)
    return matches


def _extract_columns_from_inspection(insp: FileInspection) -> List[str]:
    """Extract column/field names from any inspection result.

    Works for CSV, XLSX, Vector (SHP/GeoJSON/GPKG), NetCDF, JSON.
    Returns lowercased column names for pattern matching.
    """
    columns: List[str] = []
    fmt = (insp.format or "").lower()
    data = insp.inspection or {}

    if fmt == "csv":
        raw = data.get("columns", [])
        if isinstance(raw, list):
            columns.extend(raw)
        elif isinstance(raw, dict):
            columns.extend(raw.keys())

    elif fmt == "xlsx":
        sheets = data.get("sheets", {})
        if isinstance(sheets, dict):
            for sheet_info in sheets.values():
                if isinstance(sheet_info, dict):
                    cols = sheet_info.get("columns", [])
                    if isinstance(cols, list):
                        columns.extend(cols)
                    elif isinstance(cols, dict):
                        columns.extend(cols.keys())

    elif fmt in ("shapefile", "geojson", "geopackage", "fgdb", "vector"):
        raw = data.get("columns", [])
        if isinstance(raw, list):
            columns.extend(raw)
        # Also try col_details keys
        col_details = data.get("col_details", {})
        if isinstance(col_details, dict):
            columns.extend(col_details.keys())

    elif fmt == "netcdf":
        variables = data.get("variables", [])
        if isinstance(variables, list):
            columns.extend(variables)

    elif fmt == "json":
        fields = data.get("fields", [])
        if isinstance(fields, list):
            columns.extend(fields)
        # Also try top-level keys
        keys = data.get("keys", [])
        if isinstance(keys, list):
            columns.extend(keys)

    # Normalize: lowercase, replace spaces with underscores, deduplicate
    normalized = []
    seen = set()
    for c in columns:
        if not isinstance(c, str):
            continue
        low = c.lower().strip()
        if low and low not in seen:
            normalized.append(low)
            seen.add(low)
    return normalized


_COMP_LABEL = {"H": "Hazard", "E": "Exposure", "V": "Vulnerability", "L": "Loss"}


def _match_column_signals(
    columns: List[str],
    column_detection: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Match column names against YAML column_detection patterns.

    Returns list of dicts with signal details, sorted by weight descending.
    """
    results: List[Dict[str, Any]] = []

    for group_name, group_info in column_detection.items():
        if not isinstance(group_info, dict):
            continue

        compiled = group_info.get("patterns_compiled", [])
        if not compiled:
            continue

        component = group_info.get("component", "")
        weight_str = group_info.get("weight", "low")
        weight = _WEIGHT_MAP.get(weight_str, 1)
        label = group_info.get("label", group_name)
        sub_category = group_info.get("sub_category")

        matched_cols = []
        for col in columns:
            for pat in compiled:
                if pat.search(col):
                    matched_cols.append(col)
                    break  # one match per column is enough

        if matched_cols:
            results.append({
                "group": group_name,
                "component": component,
                "sub_category": sub_category,
                "weight": weight,
                "label": label,
                "matched_columns": matched_cols,
            })

    # Sort by weight descending
    results.sort(key=lambda x: -x["weight"])
    return results


def _maybe_split_group(
    group: FileGroup,
    rows: List[Dict],
    min_files: int = 20,
    _depth: int = 0,
    _max_depth: int = 4,
) -> List[FileGroup]:
    """Recursively split a large group by folder structure if warranted.

    Only splits groups with *min_files* or more files. Recurses up to
    *_max_depth* levels deep so that deep hierarchies (e.g., GIS/Raster_data/
    2_CurrentClimate_Baseline_DoNothing/) are properly decomposed.

    Returns the original group as a 1-element list if splitting is not useful.
    """
    if len(group.files) < min_files or _depth >= _max_depth:
        return [group]

    # Determine how many path components the group name already covers
    group_norm = group.name.replace("::", "/").replace("\\", "/")
    group_depth = len(group_norm.split("/"))

    sub_groups: Dict[str, List[str]] = {}
    for fpath in group.files:
        # Skip __MACOSX artifacts entirely
        if "__MACOSX" in fpath:
            continue
        # Normalize separators
        norm = fpath.replace("::", "/").replace("\\", "/")
        parts = norm.split("/")
        # Pick the next component after the group's current depth,
        # but ONLY if it's a directory (has more components after it).
        # If the next component is a leaf filename, keep in parent group.
        if len(parts) > group_depth + 1:
            sub_key = f"{group.name}/{parts[group_depth]}"
        else:
            sub_key = group.name
        sub_groups.setdefault(sub_key, []).append(fpath)

    # Only split if we get >1 meaningful sub-groups (not counting parent)
    real_sub = {k: v for k, v in sub_groups.items() if k != group.name}
    if len(real_sub) <= 1:
        return [group]

    # Build row lookup for size calculation
    path_size = {r["path"]: r.get("size_bytes", 0) for r in rows}

    result = []
    for key, files in sorted(sub_groups.items()):
        total_size = sum(path_size.get(f, 0) for f in files)
        fmts = sorted(set(
            Path(f.split("::")[-1] if "::" in f else f).suffix.lower().lstrip(".")
            for f in files if Path(f.split("::")[-1] if "::" in f else f).suffix
        ))
        sub = FileGroup(
            name=key, files=files, formats=fmts,
            total_size_bytes=total_size, hevl=[],
        )
        # Recurse: if a sub-group is still large, try splitting deeper
        if len(files) >= min_files:
            result.extend(_maybe_split_group(sub, rows, min_files, _depth + 1, _max_depth))
        else:
            result.append(sub)
    return result


def classify_group(group: FileGroup) -> FileGroup:
    """
    Classify a file group into HEVL components based on file names,
    folder names, and inspection results.
    """
    # Build a text blob from all file paths + group name
    text_blob = group.name + " " + " ".join(group.files)

    hevl = set()
    evidence = []
    hazard_types = set()
    exposure_categories = set()

    # Check hazard signals
    for sig_name, sig_info in HAZARD_SIGNALS.items():
        matches = _match_signals(text_blob, sig_info["patterns"])
        if matches:
            hevl.add("H")
            hazard_types.add(sig_info["hazard_type"])
            evidence.append(f"Hazard({sig_name}): matched {matches[:3]}")

    # Check exposure signals
    for sig_name, sig_info in EXPOSURE_SIGNALS.items():
        matches = _match_signals(text_blob, sig_info["patterns"])
        if matches:
            hevl.add("E")
            exposure_categories.add(sig_info["category"])
            evidence.append(f"Exposure({sig_name}): matched {matches[:3]}")

    # Check vulnerability signals
    v_matches = _match_signals(text_blob, VULNERABILITY_SIGNALS["patterns"])
    if v_matches:
        hevl.add("V")
        evidence.append(f"Vulnerability: matched {v_matches[:3]}")

    # Check loss signals (but skip for DEM/DTM-dominant groups where loss
    # keywords in READMEs can cause false positives)
    l_matches = _match_signals(text_blob, LOSS_SIGNALS["patterns"])
    if l_matches:
        hevl.add("L")
        evidence.append(f"Loss: matched {l_matches[:3]}")

    # Check DEM (supporting data — classify as hazard input, not loss)
    dem_matches = _match_signals(text_blob, DEM_SIGNALS)
    is_dem_group = False
    if dem_matches:
        is_dem_group = True
        if "H" not in hevl:
            hevl.add("H")
            hazard_types.add("dem_dtm")
        evidence.append(f"DEM/DTM detected: {dem_matches[:2]} (supporting hazard input)")

    # -- Column-level detection (v2, YAML-driven) --------------------------
    # Snapshot filename-only HEVL for conflict detection
    filename_hevl = set(hevl)

    # Extract columns from all inspections (CSV, XLSX, SHP, GeoJSON, etc.)
    all_columns: List[str] = []
    for insp in group.inspections:
        all_columns.extend(_extract_columns_from_inspection(insp))
    all_columns = list(set(all_columns))  # deduplicate

    # Match against YAML column_detection patterns
    col_signals = _match_column_signals(all_columns, COLUMN_DETECTION)

    column_evidence: List[str] = []
    conflicts: List[str] = []
    col_total_weight = 0

    for sig in col_signals:
        comp = sig["component"]
        hevl.add(comp)
        col_total_weight += sig["weight"]
        col_label = _COMP_LABEL.get(comp, comp)
        col_ev = f"{col_label}({sig['label']}): columns {sig['matched_columns'][:3]}"
        column_evidence.append(col_ev)
        evidence.append(col_ev)
        if sig.get("sub_category") and comp == "E":
            exposure_categories.add(sig["sub_category"])
        if sig.get("sub_category") and comp == "H":
            hazard_types.add(sig["sub_category"])

    # Detect conflicts (filename signals vs column signals)
    column_hevl = {s["component"] for s in col_signals}
    if column_hevl:
        for comp in filename_hevl - column_hevl:
            conflicts.append(
                f"{_COMP_LABEL.get(comp, comp)} in filename/path but no supporting columns"
            )
        for comp in column_hevl - filename_hevl:
            conflicts.append(
                f"{_COMP_LABEL.get(comp, comp)} found in columns but not in filename/path"
            )

    # Naming pattern analysis (extract scenarios, return periods, hazard subtypes)
    naming = analyze_naming_patterns(group.files)
    group_naming = naming  # store for markdown rendering
    if naming["scenarios"]:
        evidence.append(f"Scenarios: {', '.join(naming['scenarios'])}")
    if naming["return_periods"]:
        rps = ", ".join(str(r) for r in naming["return_periods"])
        evidence.append(f"Return periods: {rps} yr")
    if naming["hazard_subtypes"]:
        for hs in naming["hazard_subtypes"]:
            hazard_types.add(hs)
        evidence.append(f"Hazard subtypes: {', '.join(naming['hazard_subtypes'])}")
    if naming["gmpes"]:
        evidence.append(f"GMPEs: {', '.join(naming['gmpes'])}")
    if naming["intensity_measures"]:
        evidence.append(f"Intensity measures: {', '.join(naming['intensity_measures'])}")
    if naming["asset_types"]:
        evidence.append(f"Vulnerability asset types: {', '.join(naming['asset_types'])}")

    # Store naming analysis on group for rendering
    group._naming = naming  # type: ignore[attr-defined]

    # Post-classification: if group is DEM/DTM-only, remove false Loss
    # (DEM READMEs may mention "damage" from the broader project context)
    if is_dem_group and "L" in hevl:
        # Only keep L if there's *strong* loss evidence beyond just pattern matching
        has_strong_loss = any("column" in e.lower() or "metric" in e.lower()
                             for e in evidence if "Loss" in e)
        if not has_strong_loss:
            hevl.discard("L")
            evidence = [e for e in evidence if not e.startswith("Loss:")]

    # Confidence scoring — column evidence strengthens classification
    has_column_evidence = bool(col_signals)
    has_filename_evidence = bool(filename_hevl)
    has_conflicts = bool(conflicts)

    if col_total_weight >= 6:
        # Strong column evidence alone → high confidence
        confidence = "high"
    elif has_column_evidence and has_filename_evidence and not has_conflicts:
        # Both sources agree → high confidence
        confidence = "high"
    elif has_conflicts:
        # Disagreement between filename and columns → medium (flag for review)
        confidence = "medium"
    elif len(evidence) >= 3:
        confidence = "high"
    elif len(evidence) >= 1:
        confidence = "medium"
    else:
        confidence = "low"

    group.hevl = sorted(hevl)
    group.hazard_types = sorted(hazard_types)
    group.exposure_categories = sorted(exposure_categories)
    group.confidence = confidence
    group.evidence = evidence
    group.conflicts = conflicts
    group.column_evidence = column_evidence

    return group


# ---------------------------------------------------------------------------
# Gap analysis
# ---------------------------------------------------------------------------

def analyze_gaps(group: FileGroup) -> GapAnalysis:
    """Check what RDLS metadata we can derive vs what's missing."""
    available = {}
    missing_req = []
    missing_rec = []
    actions = []

    # What we CAN derive from the data
    if group.name:
        available["title"] = f"Derivable from folder/ZIP name: {group.name}"
    if group.hevl:
        available["risk_data_type"] = ", ".join(group.hevl)

    # Check inspections for spatial info
    has_spatial = False
    has_crs = False
    for insp in group.inspections:
        d = insp.inspection
        if "bounds" in d or "bbox" in d:
            has_spatial = True
            available["spatial.bbox"] = str(d.get("bounds") or d.get("bbox"))
        if "crs" in d or "crs_string" in d or "crs_short" in d:
            has_crs = True
            available["spatial.crs"] = d.get("crs_short") or d.get("crs") or "present"
        if "columns" in d:
            available["resource.schema"] = f"{len(d['columns'])} fields detected"

    if group.formats:
        available["resources.format"] = ", ".join(group.formats)

    # Check required fields
    if "title" not in available:
        missing_req.append("title")
    # Always missing from file inspection alone:
    missing_req.append("attributions (publisher, creator, contact_point)")
    missing_req.append("license")
    if "id" not in available:
        available["id"] = "Can be generated from naming convention"

    if not has_spatial:
        missing_req.append("spatial (bbox, country)")
        actions.append("Extract spatial extent from geospatial files")

    # Recommended
    missing_rec.append("description")
    if "H" in group.hevl and not group.hazard_types:
        missing_rec.append("hazard_type")
    if "E" in group.hevl and not group.exposure_categories:
        missing_rec.append("exposure_category")
    if "H" in group.hevl:
        missing_rec.append("process_type")
        missing_rec.append("analysis_type (probabilistic/deterministic)")
    if not has_crs:
        actions.append("Determine CRS from geospatial file headers")

    # Component-specific gaps
    if "H" in group.hevl:
        actions.append("Confirm return periods / scenario definitions from data provider")
    if "V" in group.hevl:
        actions.append("Document function_type, approach, and relationship")
    if "L" in group.hevl:
        actions.append("Identify loss metric definitions and currency")
    if "E" in group.hevl:
        actions.append("Confirm exposure taxonomy source and currency for replacement values")

    return GapAnalysis(
        group_name=group.name,
        available=available,
        missing_required=missing_req,
        missing_recommended=missing_rec,
        actions=actions,
    )


# ---------------------------------------------------------------------------
# Suggested RDLS dataset structure
# ---------------------------------------------------------------------------

def suggest_datasets(groups: List[FileGroup]) -> List[Dict[str, Any]]:
    """Suggest how to map file groups to RDLS dataset records."""
    datasets = []
    for g in groups:
        if not g.hevl:
            continue

        # Determine primary risk_data_type
        type_map = {"H": "hazard", "E": "exposure", "V": "vulnerability", "L": "loss"}
        primary_types = [type_map[h] for h in g.hevl]

        # If a group has both H and V (e.g., multi-hazard with vulnerability), split suggestion
        if len(primary_types) > 1:
            for pt in primary_types:
                datasets.append({
                    "source_group": g.name,
                    "risk_data_type": pt,
                    "hazard_types": g.hazard_types if pt == "hazard" else [],
                    "exposure_categories": g.exposure_categories if pt == "exposure" else [],
                    "formats": g.formats,
                    "file_count": len(g.files),
                    "note": f"Split from multi-component group ({', '.join(primary_types)})",
                })
        else:
            datasets.append({
                "source_group": g.name,
                "risk_data_type": primary_types[0],
                "hazard_types": g.hazard_types,
                "exposure_categories": g.exposure_categories,
                "formats": g.formats,
                "file_count": len(g.files),
            })

    return datasets


# ---------------------------------------------------------------------------
# Markdown report generator
# ---------------------------------------------------------------------------

def render_review_markdown(review: ReviewResult) -> str:
    """Generate a human-readable review markdown from structured results."""
    lines = []
    target_name = Path(review.target).name

    lines.append(f"# [Data Review] {target_name}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"| Field | Value |")
    lines.append(f"|-------|-------|")
    lines.append(f"| **Target** | `{review.target}` |")
    lines.append(f"| **Generated** | {review.generated_utc} |")
    lines.append(f"| **Files** | {review.stats.get('files', '?')} |")
    lines.append(f"| **Total size** | {review.stats.get('total_human', '?')} |")
    lines.append(f"| **Dataset groups** | {len(review.file_groups)} |")

    # HEVL summary
    all_hevl = set()
    for g in review.file_groups:
        all_hevl.update(g.hevl)
    lines.append(f"| **HEVL coverage** | {', '.join(sorted(all_hevl)) or 'none detected'} |")
    lines.append("")

    # Project metadata (from README / text files)
    if review.project_metadata:
        lines.append("## Project Metadata (extracted from README/text files)")
        lines.append("")
        lines.append("| Field | Value |")
        lines.append("|-------|-------|")
        label_map = {
            "project_title": "Project Title",
            "provider": "Data Provider",
            "financer": "Financer",
            "project_ref": "Project Reference",
            "contact": "Contact",
            "date_range": "Date Range",
        }
        for key, value in review.project_metadata.items():
            label = label_map.get(key, key)
            lines.append(f"| **{label}** | {value} |")
        lines.append("")

    # Dataset inventory table
    lines.append("## Dataset Groups")
    lines.append("")
    lines.append("| Group | Files | Size | Formats | HEVL | Confidence | Key Evidence |")
    lines.append("|-------|------:|------|---------|------|------------|-------------|")
    for g in review.file_groups:
        from .inventory import human_size
        size = human_size(g.total_size_bytes)
        fmts = ", ".join(g.formats[:4])
        hevl = "".join(g.hevl) or "?"
        ev = "; ".join(g.evidence[:2]) if g.evidence else "-"
        lines.append(f"| {g.name} | {len(g.files)} | {size} | {fmts} | {hevl} | {g.confidence} | {ev} |")
    lines.append("")

    # ZIP archive contents (tree view)
    zip_tree = build_zip_tree(review.inventory_rows)
    if zip_tree:
        lines.append("## ZIP Archive Contents")
        lines.append("")
        for zip_name, tree in sorted(zip_tree.items()):
            from .inventory import human_size
            total_files = _count_tree_files(tree)
            total_size = human_size(tree["size"] + sum(
                _sum_tree_size(c) for c in tree["children"].values()
            ))
            lines.append(f"### {zip_name} ({total_files} files, {total_size})")
            lines.append("")
            lines.append("```")
            tree_lines = _render_tree(tree)
            # Also show root-level file count if any
            if tree["files"] > 0:
                lines.append(f"  ({tree['files']} files at root level)")
            lines.extend(tree_lines)
            lines.append("```")
            lines.append("")

    # File inspections summary
    inspected = [i for i in review.inspections if i.inspection.get("error") is None]
    if inspected:
        lines.append("## File Inspections")
        lines.append("")
        for insp in inspected:
            d = insp.inspection
            lines.append(f"### `{Path(insp.path).name}` ({insp.format})")
            lines.append("")
            if "crs" in d or "crs_short" in d:
                lines.append(f"- **CRS**: {d.get('crs_short') or d.get('crs')}")
            if "bounds" in d or "bbox" in d:
                raw = d.get("bounds") or d.get("bbox")
                if isinstance(raw, (list, tuple)) and len(raw) == 4:
                    fmt = [f"{float(v):.6f}" for v in raw]
                    lines.append(f"- **Bounds**: [{', '.join(fmt)}]")
                else:
                    lines.append(f"- **Bounds**: {raw}")
            if "width" in d:
                lines.append(f"- **Size**: {d.get('width')}x{d.get('height')} pixels")
            if "pixel_scale_x" in d:
                lines.append(f"- **Resolution**: {d['pixel_scale_x']:.6f} x {d['pixel_scale_y']:.6f}")
            if "feature_count" in d:
                lines.append(f"- **Features**: {d['feature_count']:,}")
            if "geometry_type" in d:
                lines.append(f"- **Geometry**: {d['geometry_type']}")
            if "columns" in d:
                cols = d["columns"]
                if len(cols) <= 15:
                    lines.append(f"- **Fields** ({len(cols)}): {', '.join(cols)}")
                else:
                    # Multi-line for many fields
                    lines.append(f"- **Fields** ({len(cols)}):")
                    for chunk_start in range(0, len(cols), 8):
                        chunk = cols[chunk_start:chunk_start + 8]
                        lines.append(f"  {', '.join(chunk)}")

            # Band statistics (raster)
            if "band_stats" in d:
                for bs in d["band_stats"]:
                    if "note" in bs:
                        lines.append(f"- **Band {bs['band']}**: {bs['note']}")
                    else:
                        lines.append(
                            f"- **Band {bs['band']}**: min={bs['min']:.4g}, "
                            f"max={bs['max']:.4g}, mean={bs['mean']:.4g}, "
                            f"nodata={bs['nodata_pct']:.1f}%"
                        )

            # Column details (vector attribute analysis)
            if "col_details" in d:
                col_details = d["col_details"]
                # Show summary for interesting columns
                detail_lines = []
                for col_name, info in col_details.items():
                    if "unique_values" in info:
                        vals = info["unique_values"]
                        val_str = ", ".join(str(v) for v in vals[:15])
                        if len(vals) > 15:
                            val_str += ", ..."
                        detail_lines.append(
                            f"  - `{col_name}`: {info['unique_count']} unique — {val_str}"
                        )
                    elif "min" in info and "max" in info:
                        detail_lines.append(
                            f"  - `{col_name}`: range [{info['min']}, {info['max']}]"
                            f" ({info['unique_count']} unique)"
                        )
                if detail_lines:
                    lines.append("- **Column details**:")
                    lines.extend(detail_lines)

            # JSON-specific fields
            if "type" in d and insp.format == "json":
                jtype = d["type"]
                if jtype == "array":
                    lines.append(f"- **Structure**: JSON array ({d.get('count', '?')} items)")
                    if "fields" in d:
                        lines.append(f"- **Fields**: {', '.join(str(f) for f in d['fields'])}")
                elif jtype == "object":
                    if d.get("is_geojson"):
                        lines.append(f"- **Structure**: GeoJSON FeatureCollection ({d.get('feature_count', '?')} features)")
                        if "fields" in d:
                            lines.append(f"- **Properties**: {', '.join(str(f) for f in d['fields'])}")
                    elif d.get("is_columnar"):
                        lines.append(f"- **Structure**: Columnar JSON ({d.get('record_count', '?'):,} records)")
                        if "fields" in d:
                            lines.append(f"- **Fields**: {', '.join(str(f) for f in d['fields'])}")
                    else:
                        keys = d.get("keys", [])
                        lines.append(f"- **Structure**: JSON object ({len(keys)} top-level keys)")
                        if keys:
                            lines.append(f"- **Keys**: {', '.join(str(k) for k in keys)}")
            if "sheets" in d:
                for sname, sinfo in d["sheets"].items():
                    cols = sinfo.get("columns", [])
                    row_ct = sinfo.get("row_count", sinfo.get("max_row", "?"))
                    ds_cols = sinfo.get("damage_states", [])
                    uniq = sinfo.get("unique_counts", {})
                    parts = [f"**Sheet** `{sname}`: {row_ct} rows, {len(cols)} columns"]
                    if len(cols) <= 10:
                        parts.append(f"({', '.join(str(c) for c in cols)})")
                    else:
                        parts.append(f"({', '.join(str(c) for c in cols[:8])}, ...)")
                    lines.append(f"- {' '.join(parts)}")
                    if ds_cols:
                        lines.append(f"  - Damage states: {', '.join(ds_cols)}")
                    if uniq:
                        uniq_strs = [f"{k}={v}" for k, v in list(uniq.items())[:6]]
                        lines.append(f"  - Unique counts: {', '.join(uniq_strs)}")
            if "text_excerpt" in d:
                excerpt = d["text_excerpt"][:300].replace("\n", " ")
                lines.append(f"- **Text**: {excerpt}")
            lines.append("")

    # Naming analysis summary (per group)
    has_naming = any(hasattr(g, "_naming") and g._naming.get("matched_count", 0) > 0
                     for g in review.file_groups)
    if has_naming:
        lines.append("## Naming Pattern Analysis")
        lines.append("")
        for g in review.file_groups:
            naming = getattr(g, "_naming", None)
            if not naming or naming.get("matched_count", 0) == 0:
                continue
            lines.append(f"### {g.name}")
            lines.append("")
            if naming["scenarios"]:
                lines.append(f"- **Climate scenarios**: {', '.join(naming['scenarios'])}")
            if naming["return_periods"]:
                rps = ", ".join(str(r) for r in naming["return_periods"])
                lines.append(f"- **Return periods**: {rps} yr")
            if naming["hazard_subtypes"]:
                lines.append(f"- **Hazard subtypes**: {', '.join(naming['hazard_subtypes'])}")
            if naming["gmpes"]:
                lines.append(f"- **GMPEs**: {', '.join(naming['gmpes'])}")
            if naming["intensity_measures"]:
                lines.append(f"- **Intensity measures**: {', '.join(naming['intensity_measures'])}")
            if naming["asset_types"]:
                lines.append(f"- **Vulnerability asset types**: {', '.join(naming['asset_types'])}")
            lines.append(f"- _Matched {naming['matched_count']}/{naming['total_count']} files_")
            lines.append("")

    # Column evidence & conflicts (per group)
    has_col_evidence = any(g.column_evidence for g in review.file_groups)
    has_conflicts = any(g.conflicts for g in review.file_groups)
    if has_col_evidence or has_conflicts:
        lines.append("## Column-Level Classification Evidence")
        lines.append("")
        for g in review.file_groups:
            if not g.column_evidence and not g.conflicts:
                continue
            lines.append(f"### {g.name}")
            lines.append("")
            if g.column_evidence:
                lines.append("**Column signals detected:**")
                lines.append("")
                for ce in g.column_evidence:
                    lines.append(f"- {ce}")
                lines.append("")
            if g.conflicts:
                lines.append("> **Conflicts** (filename vs column evidence):")
                lines.append(">")
                for c in g.conflicts:
                    lines.append(f"> - {c}")
                lines.append("")

    # Intermediate files (excluded)
    if review.intermediate_files.get("total_excluded", 0) > 0:
        lines.append("## Intermediate Files (excluded from analysis)")
        lines.append("")
        lines.append(f"**{review.intermediate_files['total_excluded']}** files classified"
                      " as intermediate/working outputs and excluded from RDLS cataloguing.")
        lines.append("")
        lines.append("| Group | Excluded | Remaining | Reasons |")
        lines.append("|-------|--------:|----------:|---------|")
        for gname, info in review.intermediate_files.get("groups", {}).items():
            reasons = "; ".join(info["reasons"][:3])
            lines.append(f"| {gname} | {info['excluded_count']} | {info['remaining_count']} | {reasons} |")
        lines.append("")

    # Quality issues
    if review.quality_issues:
        lines.append("## Data Quality Issues")
        lines.append("")
        lines.append("| Group | Issues |")
        lines.append("|-------|--------|")
        for gname, issues in review.quality_issues.items():
            lines.append(f"| {gname} | {'; '.join(issues)} |")
        lines.append("")

    # Gap analysis
    lines.append("## Gap Analysis")
    lines.append("")
    lines.append("| Group | Missing Required | Missing Recommended | Actions |")
    lines.append("|-------|-----------------|--------------------:|---------|")
    for ga in review.gap_analyses:
        req = "; ".join(ga.missing_required[:3])
        rec = "; ".join(ga.missing_recommended[:3])
        act = "; ".join(ga.actions[:2])
        lines.append(f"| {ga.group_name} | {req} | {rec} | {act} |")
    lines.append("")

    # Suggested RDLS datasets
    if review.suggested_datasets:
        lines.append("## Suggested RDLS Datasets")
        lines.append("")
        lines.append("| Source Group | risk_data_type | hazard_types | exposure_categories | Files | Note |")
        lines.append("|-------------|---------------|-------------|--------------------:|------:|------|")
        for ds in review.suggested_datasets:
            ht = ", ".join(ds.get("hazard_types", []))
            ec = ", ".join(ds.get("exposure_categories", []))
            note = ds.get("note", "")
            lines.append(f"| {ds['source_group']} | {ds['risk_data_type']} | {ht} | {ec} | {ds['file_count']} | {note} |")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Shared inspection pipeline (used by review_folder & inspect_folder_for_llm)
# ---------------------------------------------------------------------------

@dataclass
class _PipelineResult:
    """Intermediate result from _inspect_pipeline (Steps 1-3)."""
    groups: List[FileGroup]
    inspections: List[FileInspection]
    stats: Dict[str, Any]
    rows: List[Dict]
    intermediate_summary: Dict[str, Any]


def _inspect_pipeline(
    target: Path,
    *,
    max_inspect: int = 30,
    verbose: bool = False,
) -> _PipelineResult:
    """Shared pipeline: inventory -> group -> filter intermediates -> split -> inspect.

    Steps 1-3 of the review workflow, factored out so that both
    ``review_folder()`` (with HEVL classification) and
    ``inspect_folder_for_llm()`` (without classification) can reuse them.
    """
    # Step 1: Inventory
    if verbose:
        print("[review] Phase 1: Inventorying...")
    inv_cfg = InventoryConfig(target=target, inspect_zips=True, verbose=False)
    rows, stats = scan_target(inv_cfg)
    if verbose:
        print(f"  {stats['files']} files, {stats['dirs']} dirs, {stats['zip_entries']} zip entries")

    # Step 2: Group files + filter intermediates
    if verbose:
        print("[review] Phase 2: Grouping files...")
    raw_groups = group_files(rows, target)

    intermediate_summary: Dict[str, Any] = {"groups": {}, "total_excluded": 0}
    for g in raw_groups:
        relevant, intermediate, reasons = classify_intermediate_files(g.files)
        if intermediate:
            intermediate_summary["groups"][g.name] = {
                "excluded_count": len(intermediate),
                "remaining_count": len(relevant),
                "reasons": reasons,
            }
            intermediate_summary["total_excluded"] += len(intermediate)
            g.files = relevant
            g.formats = sorted(set(
                Path(f.split("::")[-1] if "::" in f else f).suffix.lower().lstrip(".")
                for f in g.files
                if Path(f.split("::")[-1] if "::" in f else f).suffix
            ))
    if verbose and intermediate_summary["total_excluded"] > 0:
        print(f"  {intermediate_summary['total_excluded']} intermediate files excluded")

    # Split large groups by sub-folder structure
    groups: List[FileGroup] = []
    for g in raw_groups:
        if not g.files:
            continue
        groups.extend(_maybe_split_group(g, rows))
    groups = [g for g in groups if g.files]
    if verbose:
        print(f"  {len(groups)} groups identified")

    # Step 3: Inspect representative files from each group
    if verbose:
        print("[review] Phase 3: Inspecting files...")
    all_inspections: List[FileInspection] = []
    inspect_count = 0

    for group in groups:
        seen_formats: set = set()
        seen_names: set = set()
        small_group = len(group.files) <= 20
        for fpath in group.files:
            if inspect_count >= max_inspect:
                break

            full_path = target / fpath
            ext = Path(fpath).suffix.lower()
            fname = Path(fpath).name.lower()

            if _is_sidecar(Path(fpath).name):
                continue
            if "__MACOSX" in fpath or "/._{" in fpath:
                continue

            if small_group:
                if fname in seen_names:
                    continue
            else:
                if ext in seen_formats:
                    continue

            # Inspect files inside ZIPs via temp extraction
            if "::" in fpath:
                try:
                    ctx = resolve_and_open(fpath, target)
                    with ctx as temp_path:
                        insp = inspect_file(temp_path)
                        if insp is not None:
                            insp.path = fpath
                            all_inspections.append(insp)
                            group.inspections.append(insp)
                            seen_formats.add(ext)
                            seen_names.add(fname)
                            inspect_count += 1
                except Exception as e:
                    if verbose:
                        print(f"  [warn] ZIP inspect failed for {fpath}: {e}")
                continue

            if not full_path.exists():
                continue

            # Special handling for .gdb directories
            if ".gdb" in fpath:
                gdb_path = full_path
                while gdb_path.suffix != ".gdb" and gdb_path.parent != gdb_path:
                    gdb_path = gdb_path.parent
                if gdb_path.suffix == ".gdb" and str(gdb_path) not in [i.path for i in all_inspections]:
                    insp = inspect_fgdb(gdb_path)
                    all_inspections.append(insp)
                    group.inspections.append(insp)
                    seen_formats.add(".gdb")
                    seen_names.add(fname)
                    inspect_count += 1
                continue

            try:
                insp = inspect_file(full_path)
            except Exception as e:
                if verbose:
                    print(f"  [warn] Inspect failed for {fpath}: {e}")
                continue
            if insp is not None:
                all_inspections.append(insp)
                group.inspections.append(insp)
                seen_formats.add(ext)
                seen_names.add(fname)
                inspect_count += 1

    if verbose:
        print(f"  {len(all_inspections)} files inspected")

    return _PipelineResult(
        groups=groups,
        inspections=all_inspections,
        stats=stats,
        rows=rows,
        intermediate_summary=intermediate_summary,
    )


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def review_folder(
    target: str | Path,
    *,
    output_dir: str | Path | None = None,
    max_inspect: int = 30,
    verbose: bool = True,
) -> ReviewResult:
    """
    Full automated review of a delivery folder.

    Parameters
    ----------
    target : str | Path
        Folder or ZIP to review.
    output_dir : str | Path | None
        Where to write outputs. Defaults to {target}/_rdls_review.
    max_inspect : int
        Maximum number of files to individually inspect (for performance).
    verbose : bool
        Print progress messages.

    Returns
    -------
    ReviewResult
        Structured review with groups, inspections, gaps, and suggested datasets.
    """
    target = Path(target).resolve()
    if output_dir is None:
        output_dir = target / "_rdls_review"
    else:
        output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if verbose:
        print(f"[review] Target: {target}")

    # Steps 1-3: Shared pipeline (inventory -> group -> inspect)
    pipe = _inspect_pipeline(target, max_inspect=max_inspect, verbose=verbose)
    groups = pipe.groups
    all_inspections = pipe.inspections
    stats = pipe.stats
    rows = pipe.rows
    intermediate_summary = pipe.intermediate_summary

    # Step 4: Classify each group + quality checks
    if verbose:
        print("[review] Phase 4: Classifying HEVL...")
    quality_issues: Dict[str, List[str]] = {}
    for group in groups:
        classify_group(group)
        qi = check_data_quality(group)
        if qi:
            quality_issues[group.name] = qi
    classified = sum(1 for g in groups if g.hevl)
    if verbose:
        print(f"  {classified}/{len(groups)} groups classified")

    # Step 5: Gap analysis
    if verbose:
        print("[review] Phase 5: Gap analysis...")
    gap_analyses = [analyze_gaps(g) for g in groups]

    # Step 6: Suggest datasets
    suggested = suggest_datasets(groups)

    # Step 7: Extract project metadata from README/text files
    proj_meta = extract_readme_metadata(all_inspections)

    # Build result
    review = ReviewResult(
        target=str(target),
        generated_utc=iso_time(time.time()),
        stats=stats,
        file_groups=groups,
        inspections=all_inspections,
        gap_analyses=gap_analyses,
        suggested_datasets=suggested,
        quality_issues=quality_issues,
        project_metadata=proj_meta,
        intermediate_files=intermediate_summary,
        inventory_rows=rows,
    )

    # Write outputs
    stamp = time.strftime("%Y%m%dT%H%MZ", time.gmtime())

    # Structured JSON
    json_path = output_dir / f"review_{stamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(_serialize_review(review), f, indent=2, default=str)
    if verbose:
        print(f"  -> {json_path}")

    # Markdown report
    md = render_review_markdown(review)
    md_path = output_dir / f"review_{stamp}.md"
    md_path.write_text(md, encoding="utf-8")
    if verbose:
        print(f"  -> {md_path}")

    print(f"\n[done] Review complete: {len(groups)} groups, {len(all_inspections)} inspections, {len(suggested)} suggested datasets")

    return review


def _serialize_review(review: ReviewResult) -> Dict:
    """Convert ReviewResult to JSON-serializable dict."""
    return {
        "target": review.target,
        "generated_utc": review.generated_utc,
        "stats": review.stats,
        "file_groups": [
            {
                "name": g.name,
                "file_count": len(g.files),
                "formats": g.formats,
                "total_size_bytes": g.total_size_bytes,
                "hevl": g.hevl,
                "hazard_types": g.hazard_types,
                "exposure_categories": g.exposure_categories,
                "confidence": g.confidence,
                "evidence": g.evidence,
            }
            for g in review.file_groups
        ],
        "inspections": [
            {"path": i.path, "format": i.format, "inspection": i.inspection}
            for i in review.inspections
        ],
        "gap_analyses": [asdict(ga) for ga in review.gap_analyses],
        "suggested_datasets": review.suggested_datasets,
        "quality_issues": review.quality_issues,
        "project_metadata": review.project_metadata,
        "intermediate_files": review.intermediate_files,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="review",
        description="Automated data review for RDLS metadata preparation.",
    )
    p.add_argument("target", type=Path, help="Folder or ZIP to review")
    p.add_argument("-o", "--output-dir", type=Path, default=None,
                   help="Output directory (default: TARGET/_rdls_review)")
    p.add_argument("--max-inspect", type=int, default=30,
                   help="Max files to inspect (default: 30)")
    p.add_argument("-q", "--quiet", action="store_true", help="Suppress progress")
    return p


def main(argv: List[str] | None = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)
    review_folder(
        target=args.target,
        output_dir=args.output_dir,
        max_inspect=args.max_inspect,
        verbose=not args.quiet,
    )


if __name__ == "__main__":
    main()
